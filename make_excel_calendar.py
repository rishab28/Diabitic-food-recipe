import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_calendar():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "30-Day Tiffin Calendar"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles definition
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10, color="000000")
    bold_data_font = Font(name="Arial", size=10, bold=True, color="000000")
    tip_font = Font(name="Arial", size=9, italic=True, color="555555")
    
    title_fill = PatternFill(start_color="FF7020", end_color="FF7020", fill_type="solid")
    header_fill = PatternFill(start_color="FF7020", end_color="FF7020", fill_type="solid")
    accent_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Title Block
    ws.merge_cells("A1:G1")
    ws["A1"] = "THE 30-DAY HAPPY TIFFIN CALENDAR (BY PRIYA & KHADUFARM)"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Subtitle Block
    ws.merge_cells("A2:G2")
    ws["A2"] = "Daily Balanced Lunchbox Combos featuring pure, wax-free Himalayan ingredients from KhaduFarm."
    ws["A2"].font = Font(name="Arial", size=11, italic=True, color="000000")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Headers
    headers = [
        "Week", "Day", "Main Tiffin Item", "Side / Fruit Snack (KhaduFarm Recommended)",
        "Hidden Veggie / Swap Detail", "KhaduFarm Product Integration", "Completed? (Write Yes/No)"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    # Data Row Definitions
    data = [
        # Week 1
        ("Week 1", "Day 1 (Mon)", "Oats-Ragi Chocolate Pancakes", "KhaduFarm Apple Slices (Wax-free)", "Hidden steamed spinach in the batter", "Drizzle with KhaduFarm Raw Forest Honey", ""),
        ("Week 1", "Day 2 (Tue)", "Hidden-Veggie Cheese Rolls", "KhaduFarm Dried Apricots & Walnuts", "Finely grated carrots & paneer in soft wrap", "Pair with KhaduFarm Roasted Walnuts", ""),
        ("Week 1", "Day 3 (Wed)", "Golden Smiley Mini-Idlis", "KhaduFarm Kiwi Slices", "Pureed yellow pumpkin folded in batter", "Grease idli molds with KhaduFarm A2 Ghee", ""),
        ("Week 1", "Day 4 (Thu)", "Oats & Beetroot Chilla", "KhaduFarm Fresh Pears", "Steamed beetroot puree for pink color", "Cook chilla using pure KhaduFarm A2 Ghee", ""),
        ("Week 1", "Day 5 (Fri)", "Beetroot Pink Sauce Pasta", "KhaduFarm Plums", "Steamed beetroot blended in tomato base", "Serve with sliced KhaduFarm apples", ""),
        # Week 2
        ("Week 2", "Day 6 (Mon)", "Ragi-Date Muffins", "KhaduFarm Dried Plums (Prunes)", "Seedless date puree instead of white sugar", "Made with KhaduFarm Himalayan Plums", ""),
        ("Week 2", "Day 7 (Tue)", "Paneer & Pea Pocket Parathas", "KhaduFarm Walnut Halves", "Finely mashed green peas in paneer dough", "Roast parathas in aromatic KhaduFarm Ghee", ""),
        ("Week 2", "Day 8 (Wed)", "Tri-Color Dhokla Shapes", "KhaduFarm Persimmon slices", "Spinach (green) & Carrot (orange) juices", "Serve with a side of KhaduFarm fresh fruits", ""),
        ("Week 2", "Day 9 (Thu)", "Vegetable Semolina Upma", "KhaduFarm Dried Apricots", "Carrots & green beans cut extra fine", "Lightly roast upma suji in KhaduFarm Ghee", ""),
        ("Week 2", "Day 10 (Fri)", "Suji Mini-Pizza Discs", "KhaduFarm Apple Cubes", "Steamed pureed pumpkin in tomato pizza sauce", "Brush crust with KhaduFarm A2 Ghee", ""),
        # Week 3
        ("Week 3", "Day 11 (Mon)", "Banana-Oat Waffles", "KhaduFarm Kiwi & Plum mix", "Ripened banana sweetening, egg-free batter", "Drizzle with pure KhaduFarm Raw Honey", ""),
        ("Week 3", "Day 12 (Tue)", "Green Moong Dal Wraps", "KhaduFarm Soaked Walnuts", "Pureed spinach and coriander in dal wrap", "Fill with paneer cooked in KhaduFarm Ghee", ""),
        ("Week 3", "Day 13 (Wed)", "Mini Beetroot Tikki/Cutlets", "KhaduFarm Pear Slices", "Steamed beetroot and sweet potato mash", "Shallow fry in pure KhaduFarm Ghee", ""),
        ("Week 3", "Day 14 (Thu)", "Vegetable Oats Porridge", "KhaduFarm Dried Plums", "Finely grated bottle gourd cooked with milk/oats", "Sweeten with KhaduFarm Raw Honey", ""),
        ("Week 3", "Day 15 (Fri)", "Suji Veggie Hakka Noodles", "KhaduFarm Fresh Apple slices", "Spiralized zucchini noodles mixed with suji", "Toss noodles with sesame oil & garlic", ""),
        # Week 4
        ("Week 4", "Day 16 (Mon)", "Sweet Potato Pancakes", "KhaduFarm Dried Apricots", "Steamed sweet potato mash in pancake batter", "Drizzle with KhaduFarm Raw Honey", ""),
        ("Week 4", "Day 17 (Tue)", "Hummus & Carrot/Cucumber Sticks", "KhaduFarm Walnut Halves", "Chickpea paste with sesame and olive oil", "Excellent high-fiber protein snack combo", ""),
        ("Week 4", "Day 18 (Wed)", "Mini Rava Appe (Appam)", "KhaduFarm Persimmon Cubes", "Grated carrots and onions mixed in yogurt", "Brush appam cavity with KhaduFarm Ghee", ""),
        ("Week 4", "Day 19 (Thu)", "Masala Oats Idli", "KhaduFarm Fresh Kiwi slices", "Powdered roasted oats and steamed beans", "Grease idli plates with KhaduFarm Ghee", ""),
        ("Week 4", "Day 20 (Fri)", "Whole Wheat Burger Slider", "KhaduFarm Apple Cubes", "Potato-Paneer patty loaded with hidden spinach", "Pan-sear patty in KhaduFarm A2 Ghee", ""),
    ]
    
    current_row = 4
    for r_idx, row_data in enumerate(data, 4):
        ws.row_dimensions[r_idx].height = 28
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = data_font
            cell.border = thin_border
            
            # Formatting specifics
            if c_idx in [1, 2, 7]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            if c_idx == 1:
                cell.font = bold_data_font
                
            # Zebra striping
            if r_idx % 2 == 0:
                cell.fill = zebra_fill
                
            # Special highlighting for KhaduFarm columns
            if c_idx in [4, 6]:
                cell.fill = accent_fill
                cell.font = bold_data_font if c_idx == 6 else data_font

    # Set column widths dynamically with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            ws.column_dimensions[col_letter].width = 12
        elif col[0].column == 2:
            ws.column_dimensions[col_letter].width = 15
        elif col[0].column == 3:
            ws.column_dimensions[col_letter].width = 30
        elif col[0].column == 4:
            ws.column_dimensions[col_letter].width = 35
        elif col[0].column == 5:
            ws.column_dimensions[col_letter].width = 35
        elif col[0].column == 6:
            ws.column_dimensions[col_letter].width = 35
        elif col[0].column == 7:
            ws.column_dimensions[col_letter].width = 18

    # Save to deliverables directory
    filename = "/Users/rishabsayrta/Downloads/Diabetes and healthy Food/kids/deliverables/Empty_Tiffin_Calendar.xlsx"
    wb.save(filename)
    print(f"Excel Calendar successfully created at {filename}")

if __name__ == "__main__":
    create_calendar()
